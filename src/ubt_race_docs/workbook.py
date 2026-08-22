"""Эксель для расчёта призовых.

Судья вставляет протокол категории (номер, участник, результат) в жёлтую зону,
всё остальное считают формулы прямо в книге — пересчёт мгновенный, ничего
запускать не нужно. Правила распределения те же, что в `prizes`.

Порог, от которого считаются выигранные секунды, книга подбирает сама:
сначала по накопленным временам находится порог, при котором выплаты без
округления дают ровно фонд, а потом рядом с ним перебирается двести значений
с шагом в десятую долю секунды и берётся самое большое, при котором
округлённые до тысячи выплаты ещё умещаются в фонд.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from .race import CATEGORIES, RACE

DEFAULT_ROWS = 200
HEADER_ROW = 10
FIRST_DATA_ROW = HEADER_ROW + 1

# Подписи параметров длинные, поэтому занимают слитые A:C, а значения лежат
# в D — иначе узкая колонка «Место» обрезала бы подпись прямо в шапке.
ENTRY_FEE_CELL = "D4"
ENTRIES_CELL = "D5"
FUND_CELL = "D6"
SECOND_PRICE_CELL = "D7"
ROUNDING_CELL = "D8"
PAID_CELL = "I4"
REMAINDER_CELL = "I5"
WINNERS_CELL = "I6"
THRESHOLD_CELL = "I7"
WARNING_CELL = "A9"

# Служебные ячейки подбора порога — за краем видимой части листа.
EVEN_THRESHOLD_CELL = "V4"
BASE_PAID_CELL = "V5"
BASE_REMAINDER_CELL = "V6"
BASE_WINNERS_CELL = "V7"


def absolute(cell: str) -> str:
    """«D6» → «$D$6»: ссылка, которая не поедет при копировании ячеек."""
    column = cell[0]
    return f"${column}${cell[1:]}"


# Как разбирается вставленный результат (колонка «Время, с»).
#
# Разбор текстовый, по двоеточиям и точке, а не через VALUE: Apple Numbers,
# в отличие от Excel и LibreOffice, не читает этой функцией ни «0:34:12»,
# ни «26.1» — в русской локали у него десятичная запятая. Поэтому каждая
# часть берётся отдельно и складывается числами.
#
#   «00:29:26.1» — часы, минуты, секунды с десятыми;
#   «34:12»      — минуты и секунды;
#   «2052»       — уже секунды;
#   число        — Excel сам понял ячейку как время: доля суток → ×86400,
#                  а «34:12», прочитанное им как 34 часа 12 минут, — ×1440.
#
# Настоящий результат на 25 км в диапазон 0.5…100 попасть не может,
# поэтому числовая развилка однозначная.
MONEY_FORMAT = "# ##0"
SECONDS_FORMAT = "# ##0.#"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
PARAMETER_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14)
HINT_FONT = Font(size=9, color="808080")
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GRID_POINTS = 200
"""Сколько порогов перебирается вокруг найденного — по 0.1 секунды."""

GRID_HALF_WIDTH = 10.0
"""На сколько секунд перебор уходит в каждую сторону."""


@dataclass(frozen=True, slots=True)
class Column:
    """Колонка таблицы результатов."""

    letter: str
    header: str
    width: float
    number_format: str | None = None
    formula: str | None = None
    """Шаблон формулы; `{row}` подставляется номером строки, `{last}` — последней."""

    editable: bool = False


COLUMNS: tuple[Column, ...] = (
    # Место считается по заполненным результатам, а не по номеру строки:
    # у сошедшего результата нет, и места ему тоже не полагается.
    Column("A", "Место", 7, SECONDS_FORMAT, '=IF($E{row}="","",COUNT($E${first}:$E{row}))'),
    Column("B", "Номер", 8, editable=True),
    Column("C", "Участник", 30, editable=True),
    Column("D", "Результат", 13, editable=True),
    Column(
        "E",
        "Время, с",
        10,
        SECONDS_FORMAT,
        '=IF($D{row}="","",IFERROR('
        "IF($I{row}=-1,"
        "IF($D{row}<0.5,$D{row}*86400,IF($D{row}<100,$D{row}*1440,$D{row})),"
        "IF($I{row}=2,VALUE(LEFT($D{row},$J{row}-1))*3600"
        "+VALUE(MID($D{row},$J{row}+1,$K{row}-$J{row}-1))*60,"
        "IF($I{row}=1,VALUE(LEFT($D{row},$J{row}-1))*60,0))"
        "+IF($M{row}=0,VALUE($L{row}),"
        "VALUE(LEFT($L{row},$M{row}-1))"
        "+VALUE(MID($L{row},$M{row}+1,10))/10^(LEN($L{row})-$M{row}))"
        '),""))',
    ),
    Column(
        "F",
        "Опережение порога, с",
        13,
        SECONDS_FORMAT,
        '=IF($E{row}="","",MAX(0,{threshold}-$E{row}))',
    ),
    Column(
        "G",
        "Приз до округления, ₸",
        15,
        MONEY_FORMAT,
        '=IF($E{row}="","",$F{row}*{price})',
    ),
    # Тысяча нераспределённого остатка достаётся первому, кому ничего не досталось.
    Column(
        "H",
        "К выдаче, ₸",
        12,
        MONEY_FORMAT,
        '=IF($E{row}="","",$T{row}+IF(AND({base_remainder}>={rounding},'
        "COUNT($E${first}:$E{row})={base_winners}+1),{rounding},0))",
    ),
    Column(
        "I",
        "служебное: двоеточий",
        11,
        None,
        '=IF($D{row}="","",IF(ISTEXT($D{row}),LEN($D{row})-LEN(SUBSTITUTE($D{row},":","")),-1))',
    ),
    Column("J", "служебное: первое двоеточие", 11, None, '=IFERROR(FIND(":",$D{row}),0)'),
    Column(
        "K",
        "служебное: второе двоеточие",
        11,
        None,
        '=IFERROR(FIND(":",$D{row},$J{row}+1),0)',
    ),
    Column(
        "L",
        "служебное: секунды текстом",
        13,
        None,
        '=IF($D{row}="","",IF($I{row}=2,MID($D{row},$K{row}+1,10),'
        'IF($I{row}=1,MID($D{row},$J{row}+1,10),IF($I{row}=0,$D{row},""))))',
    ),
    Column(
        "M",
        "служебное: дробная часть",
        13,
        None,
        '=IF($L{row}="",0,IFERROR(FIND(".",SUBSTITUTE($L{row},",",".")),0))',
    ),
    # Пустым строкам подставляем заведомо недостижимое время: так формулы
    # перебора порога остаются числовыми и не спотыкаются о пустые ячейки.
    Column("N", "служебное: время для расчёта", 13, None, '=IF($E{row}="",10^9,$E{row})'),
    Column(
        "O",
        "служебное: сумма времён",
        13,
        None,
        '=IF($E{row}="","",SUM($E${first}:$E{row}))',
    ),
    # Если призовые получают ровно эти участники, то вот порог, при котором
    # выплаты без округления дают ровно фонд.
    Column(
        "P",
        "служебное: порог для места",
        14,
        SECONDS_FORMAT,
        '=IF($E{row}="","",({fund}/{price}+$O{row})/COUNT($E${first}:$E{row}))',
    ),
    Column(
        "Q",
        "служебное: порог подходит",
        13,
        None,
        '=IF($E{row}="",0,IF(AND($E{row}<$P{row},OR($E{next}="",$P{row}<=$E{next})),1,0))',
    ),
    Column(
        "R",
        "служебное: перебор порога",
        14,
        SECONDS_FORMAT,
        '=IF({even_threshold}=0,"",ROUND({even_threshold},1)-'
        + str(GRID_HALF_WIDTH)
        + "+({row}-{first})*"
        + str(round(2 * GRID_HALF_WIDTH / GRID_POINTS, 3))
        + ")",
    ),
    Column(
        "S",
        "служебное: выплаты при пороге",
        16,
        MONEY_FORMAT,
        '=IF($R{row}="","",SUMPRODUCT(($N${first}:$N${last}<$R{row})*'
        "ROUND(($R{row}-$N${first}:$N${last})/({rounding}/{price}),0))*{rounding})",
    ),
    Column(
        "T",
        "служебное: приз без добавки",
        16,
        MONEY_FORMAT,
        '=IF($E{row}="","",ROUND($G{row}/{rounding},0)*{rounding})',
    ),
)


PARAMETERS: tuple[tuple[str, str], ...] = (
    (ENTRY_FEE_CELL, "Стартовый взнос, ₸"),
    (ENTRIES_CELL, "Оплатило взносов, чел."),
    (FUND_CELL, "Призовой фонд, ₸"),
    (SECOND_PRICE_CELL, "Цена секунды, ₸"),
    (ROUNDING_CELL, "Округление выплаты, ₸"),
)

TOTALS: tuple[tuple[str, str], ...] = (
    (PAID_CELL, "Выплачено, ₸"),
    (REMAINDER_CELL, "Остаток фонда, ₸"),
    (WINNERS_CELL, "Призёров"),
    (THRESHOLD_CELL, "Пороговое время, с"),
)

SERVICE_CELLS: tuple[tuple[str, str], ...] = (
    (EVEN_THRESHOLD_CELL, "служебное: порог без округления"),
    (BASE_PAID_CELL, "служебное: выплаты без добавки"),
    (BASE_REMAINDER_CELL, "служебное: остаток без добавки"),
    (BASE_WINNERS_CELL, "служебное: призёров без добавки"),
)

OUT_OF_ORDER_WARNING = (
    "Внимание: протокол вставлен не по возрастанию времени — "
    "проверьте порядок строк, призовые посчитаны неверно"
)

INSTRUCTIONS: tuple[tuple[str, bool], ...] = (
    ("Как считать призовые", True),
    ("", False),
    ("1. Откройте лист нужной категории — «Мужчины» или «Женщины». Фонды раздельные:", False),
    ("   взносы женщин разыгрываются среди женщин, взносы мужчин — среди мужчин.", False),
    ("2. Вставьте протокол в жёлтые колонки «Номер», «Участник», «Результат»,", False),
    ("   начиная со строки 11, в порядке финиша — от лучшего времени к худшему.", False),
    ("   Сошедших и снятых оставляйте без результата — место и приз им", False),
    ("   не посчитаются, а во взносах они учтутся.", False),
    ("3. Результат понимается в том виде, в каком его печатает протокол:", False),
    ("   00:29:26.1, 34:12 или просто число секунд. Десятые доли учитываются.", False),
    ("4. «Оплатило взносов» считается по колонке «Участник»: взнос платит и тот,", False),
    ("   кто потом сошёл или был снят. Число можно перебить руками.", False),
    ("5. Колонка «К выдаче» — это то, что вручается наличными.", False),
    ("6. Если строки перепутаны местами, над таблицей загорится красное", False),
    ("   предупреждение — призовые в этом случае считать нельзя.", False),
    ("", False),
    ("Правило из положения", True),
    ("", False),
    ("Каждая выигранная секунда стоит 100 ₸. Считаются они от порогового времени:", False),
    ("участник получает за столько секунд, на сколько опередил порог. Победитель", False),
    ("тем самым получает за отрыв от второго, первые двое — за отрыв от третьего", False),
    ("и так далее, как и написано в положении.", False),
    ("", False),
    ("Само пороговое время книга подбирает так, чтобы фонд разошёлся целиком:", False),
    ("суммы округляются до ближайшей тысячи, а порог двигается с шагом в десятую", False),
    ("долю секунды, пока сумма выплат не сравняется с фондом. Найденный порог", False),
    ("показан в шапке — по нему легко проверить любую строку вручную.", False),
    ("", False),
    ("Если попасть в фонд точно нельзя — так бывает, когда несколько участников", False),
    ("показали совпадающее до десятой доли время, — лишняя тысяча достаётся первому", False),
    ("из тех, кому ничего не досталось, а всё сверх неё остаётся в фонде.", False),
    ("", False),
    ("Колонки после «Время, с» — расчётные, их менять не нужно. Служебные колонки", False),
    ("справа разбирают вставленный результат и перебирают пороговое время.", False),
)


def _write_header(sheet: Worksheet, title: str, last_row: int) -> None:
    sheet["A1"] = title
    sheet["A1"].font = TITLE_FONT
    sheet["A2"] = (
        "Вставьте протокол в жёлтые колонки начиная со строки 11 — остальное посчитается само."
    )
    sheet["A2"].font = HINT_FONT

    first = FIRST_DATA_ROW
    fund = absolute(FUND_CELL)
    values = {
        ENTRY_FEE_CELL: RACE.prizes.entry_fee,
        ENTRIES_CELL: f"=COUNTA($C${first}:$C${last_row})",
        FUND_CELL: f"={absolute(ENTRY_FEE_CELL)}*{absolute(ENTRIES_CELL)}",
        SECOND_PRICE_CELL: RACE.prizes.tenge_per_second,
        ROUNDING_CELL: RACE.prizes.payout_step,
        PAID_CELL: f"=SUM($H${first}:$H${last_row})",
        REMAINDER_CELL: f"={fund}-{absolute(PAID_CELL)}",
        WINNERS_CELL: f'=COUNTIF($H${first}:$H${last_row},">0")',
        # Из перебора берём самый большой порог, при котором выплаты ещё
        # умещаются в фонд: сумма выплат по порогу не убывает.
        THRESHOLD_CELL: (
            f"=SUMPRODUCT(MAX(($S${first}:$S${last_row}<={fund})"
            f'*($R${first}:$R${last_row}<>"")*$R${first}:$R${last_row}))'
        ),
        EVEN_THRESHOLD_CELL: f"=SUMIF($Q${first}:$Q${last_row},1,$P${first}:$P${last_row})",
        BASE_PAID_CELL: f"=SUM($T${first}:$T${last_row})",
        BASE_REMAINDER_CELL: f"={fund}-{absolute(BASE_PAID_CELL)}",
        BASE_WINNERS_CELL: f'=COUNTIF($T${first}:$T${last_row},">0")',
    }

    for cells, label_span in ((PARAMETERS, ("A", "C")), (TOTALS, ("F", "H"))):
        for value_cell, label in cells:
            row = value_cell[1:]
            left, right = label_span
            # Подпись занимает несколько колонок: узкая «Место» её бы обрезала.
            sheet.merge_cells(f"{left}{row}:{right}{row}")
            label_cell = sheet[f"{left}{row}"]
            label_cell.value = label
            label_cell.font = Font(bold=True, size=10)
            label_cell.alignment = Alignment(horizontal="left", vertical="center")

            cell = sheet[value_cell]
            cell.value = values[value_cell]
            cell.number_format = SECONDS_FORMAT if value_cell == THRESHOLD_CELL else MONEY_FORMAT
            cell.fill = PARAMETER_FILL
            cell.border = CELL_BORDER

    for service_cell, label in SERVICE_CELLS:
        row = service_cell[1:]
        sheet[f"U{row}"] = label
        sheet[f"U{row}"].font = HINT_FONT
        sheet[service_cell] = values[service_cell]
        sheet[service_cell].number_format = SECONDS_FORMAT

    warning = sheet[WARNING_CELL]
    warning.value = (
        f"=IF(SUMPRODUCT(($N${first}:$N${last_row - 1}<10^9)"
        f"*($N${first + 1}:$N${last_row}<10^9)"
        f"*($N${first}:$N${last_row - 1}>$N${first + 1}:$N${last_row}))>0,"
        f'"{OUT_OF_ORDER_WARNING}","")'
    )
    warning.font = Font(bold=True, size=10, color="C00000")


def _write_table(sheet: Worksheet, last_row: int) -> None:
    for column in COLUMNS:
        cell = sheet[f"{column.letter}{HEADER_ROW}"]
        cell.value = column.header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.column_dimensions[column.letter].width = column.width

    sheet.row_dimensions[HEADER_ROW].height = 32

    for row in range(FIRST_DATA_ROW, last_row + 1):
        for column in COLUMNS:
            cell = sheet[f"{column.letter}{row}"]
            if column.formula is not None:
                cell.value = column.formula.format(
                    row=row,
                    next=row + 1,
                    first=FIRST_DATA_ROW,
                    last=last_row,
                    header=HEADER_ROW,
                    fund=absolute(FUND_CELL),
                    price=absolute(SECOND_PRICE_CELL),
                    rounding=absolute(ROUNDING_CELL),
                    threshold=absolute(THRESHOLD_CELL),
                    even_threshold=absolute(EVEN_THRESHOLD_CELL),
                    base_remainder=absolute(BASE_REMAINDER_CELL),
                    base_winners=absolute(BASE_WINNERS_CELL),
                )
            if column.number_format is not None:
                cell.number_format = column.number_format
            if column.editable:
                cell.fill = INPUT_FILL
            cell.border = CELL_BORDER

    sheet[f"K{HEADER_ROW}"].font = Font(bold=True, color="FFFFFF", size=10)
    sheet.freeze_panes = f"A{FIRST_DATA_ROW}"


def _write_instructions(sheet: Worksheet) -> None:
    sheet.column_dimensions["A"].width = 100
    for index, (line, is_heading) in enumerate(INSTRUCTIONS, start=1):
        cell = sheet[f"A{index}"]
        cell.value = line
        cell.font = Font(bold=True, size=12) if is_heading else Font(size=11)


def build_workbook(output: Path, rows: int = DEFAULT_ROWS) -> Path:
    """Собрать книгу с листом на каждую категорию и инструкцией."""
    if rows < 1:
        raise ValueError(f"в таблице должна быть хотя бы одна строка, получено {rows}")

    last_row = HEADER_ROW + rows
    book = Workbook()
    book.remove(book.active)

    for category in CATEGORIES:
        sheet = book.create_sheet(category.name.ru)
        _write_header(sheet, f"{category.name.one_line()} — призовые", last_row)
        _write_table(sheet, last_row)

    _write_instructions(book.create_sheet("Инструкция"))

    book.properties.title = f"Призовые · {RACE.short_title}"
    book.properties.creator = RACE.organizer
    book.properties.subject = RACE.title.ru

    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)
    return output
