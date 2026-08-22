"""Проверки книги расчёта призовых.

Формулы в книге считает не Python, поэтому кроме структуры мы прогоняем её
через LibreOffice и сверяем с эталонным расчётом из `prizes` — если он есть
в системе.
"""

from __future__ import annotations

import csv
import shutil
import subprocess
from collections.abc import Callable
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ubt_race_docs.prizes import Result, distribute, fund_from_entries
from ubt_race_docs.race import RACE
from ubt_race_docs.workbook import (
    ENTRIES_CELL,
    ENTRY_FEE_CELL,
    FIRST_DATA_ROW,
    FUND_CELL,
    HEADER_ROW,
    OUT_OF_ORDER_WARNING,
    PAID_CELL,
    REMAINDER_CELL,
    ROUNDING_CELL,
    SECOND_PRICE_CELL,
    WARNING_CELL,
    build_workbook,
)

PROTOCOL: tuple[tuple[str, int], ...] = (
    ("Гонщик 1", 2052),
    ("Гонщик 2", 2114),
    ("Гонщик 3", 2145),
    ("Гонщик 4", 2160),
    ("Гонщик 5", 2160),
    ("Гонщик 6", 2400),
    ("Гонщик 7", 2401),
)
ENTRIES = 60


@pytest.fixture(scope="module")
def workbook_path(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return build_workbook(tmp_path_factory.mktemp("workbook") / "prizes.xlsx", rows=50)


def test_one_sheet_per_category_plus_instructions(workbook_path: Path) -> None:
    book = load_workbook(workbook_path)
    assert book.sheetnames == ["Мужчины", "Женщины", "Инструкция"]


def test_parameters_come_from_the_regulations(workbook_path: Path) -> None:
    sheet = load_workbook(workbook_path)["Мужчины"]
    assert sheet[ENTRY_FEE_CELL].value == RACE.prizes.entry_fee
    assert sheet[SECOND_PRICE_CELL].value == RACE.prizes.tenge_per_second
    assert sheet[ROUNDING_CELL].value == RACE.prizes.payout_step
    assert sheet[FUND_CELL].value == "=$D$4*$D$5"


def test_parameter_labels_are_not_squeezed_into_a_narrow_column(workbook_path: Path) -> None:
    sheet = load_workbook(workbook_path)["Мужчины"]
    merges = {str(merged) for merged in sheet.merged_cells.ranges}
    for value_cell, label in ((ENTRY_FEE_CELL, "Стартовый взнос, ₸"),):
        row = value_cell[1:]
        assert f"A{row}:C{row}" in merges
        assert sheet[f"A{row}"].value == label
    span = sum(sheet.column_dimensions[letter].width for letter in "ABC")
    assert span > max(len(label) for _, label in ((ENTRY_FEE_CELL, "Оплатило взносов, чел."),))


def test_protocol_columns_are_marked_as_input(workbook_path: Path) -> None:
    sheet = load_workbook(workbook_path)["Мужчины"]
    headers = {sheet.cell(HEADER_ROW, index).value for index in range(1, 13)}
    assert {"Место", "Номер", "Участник", "Результат", "К выдаче, ₸"} <= headers
    for letter in ("B", "C", "D"):
        assert sheet[f"{letter}{FIRST_DATA_ROW}"].fill.fgColor.rgb.endswith("FFF2CC")
    assert sheet[f"E{FIRST_DATA_ROW}"].value.startswith("=")


def test_every_row_of_the_table_has_formulas(workbook_path: Path) -> None:
    sheet = load_workbook(workbook_path)["Мужчины"]
    for row in (FIRST_DATA_ROW, FIRST_DATA_ROW + 25, HEADER_ROW + 50):
        assert sheet[f"K{row}"].value == f'=IF($J{row}="","",FLOOR(MAX($J{row},0),$D$8))'
        assert "SUMIFS" in sheet[f"J{row}"].value


def test_entry_count_comes_from_the_names_not_the_times(workbook_path: Path) -> None:
    # Взнос платит и тот, кто потом сошёл: у него есть фамилия, но нет результата.
    sheet = load_workbook(workbook_path)["Мужчины"]
    assert sheet[ENTRIES_CELL].value.startswith("=COUNTA($C$")


def test_time_is_parsed_by_hand_not_by_value(workbook_path: Path) -> None:
    # Apple Numbers не умеет VALUE("0:34:12") — там таблица молча оставалась
    # пустой, — поэтому текст разбирается по двоеточиям.
    sheet = load_workbook(workbook_path)["Мужчины"]
    formula = sheet[f"E{FIRST_DATA_ROW}"].value
    assert "LEFT(" in formula and "MID(" in formula
    assert f"VALUE($D{FIRST_DATA_ROW})*86400" not in formula


def test_place_is_counted_by_results_not_by_row(workbook_path: Path) -> None:
    sheet = load_workbook(workbook_path)["Мужчины"]
    assert sheet[f"A{FIRST_DATA_ROW}"].value.startswith('=IF($E11="","",COUNT($E$11:$E11')


def test_row_count_is_validated(tmp_path: Path) -> None:
    with pytest.raises(ValueError, match="хотя бы одна строка"):
        build_workbook(tmp_path / "prizes.xlsx", rows=0)


def fill_protocol(path: Path, render_time: Callable[[int], str | float]) -> None:
    """Вставить протокол так, как это сделает судья."""
    book = load_workbook(path)
    sheet = book["Мужчины"]
    for index, (name, seconds) in enumerate(PROTOCOL):
        row = FIRST_DATA_ROW + index
        sheet[f"B{row}"] = 100 + index
        sheet[f"C{row}"] = name
        sheet[f"D{row}"] = render_time(seconds)
    sheet[ENTRIES_CELL] = ENTRIES
    book.save(path)


def recalculate(path: Path, tmp_path: Path) -> list[list[str]]:
    """Пересчитать книгу LibreOffice и вернуть лист «Мужчины» как таблицу."""
    profile = tmp_path / "profile"
    outdir = tmp_path / "csv"
    subprocess.run(
        [
            "soffice",
            f"-env:UserInstallation=file://{profile}",
            "--headless",
            "--convert-to",
            "csv:Text - txt - csv (StarCalc):44,34,76,1,,0,false,true,true,false,false,-1",
            "--outdir",
            str(outdir),
            str(path),
        ],
        check=True,
        capture_output=True,
        timeout=300,
    )
    exported = outdir / f"{path.stem}-Мужчины.csv"
    with exported.open(encoding="utf-8", newline="") as handle:
        return list(csv.reader(handle))


def cell(table: list[list[str]], address: str) -> str:
    """Значение ячейки по её адресу вида «I4» в выгруженной таблице."""
    column = ord(address[0]) - ord("A")
    return table[int(address[1:]) - 1][column]


def money(text: str) -> int:
    """Число из ячейки: LibreOffice печатает разряды через пробел."""
    return int(text.replace("\xa0", "").replace(" ", "") or 0)


TIME_FORMATS: dict[str, Callable[[int], str | float]] = {
    # Как судья может вставить одно и то же время.
    "чч:мм:сс": lambda s: f"{s // 3600}:{s % 3600 // 60:02d}:{s % 60:02d}",
    "мм:сс": lambda s: f"{s // 60}:{s % 60:02d}",
    "секунды": lambda s: s,
    "доля суток": lambda s: s / 86400,
    # Так «34:12» выглядит после того, как Excel прочитал его как 34 ч 12 мин.
    "мм:сс числом": lambda s: s / 1440,
}


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice не установлен")
@pytest.mark.parametrize("time_format", list(TIME_FORMATS), ids=list(TIME_FORMATS))
def test_formulas_match_the_reference_calculation(
    workbook_path: Path, tmp_path: Path, time_format: str
) -> None:
    path = tmp_path / "filled.xlsx"
    shutil.copy(workbook_path, path)
    fill_protocol(path, TIME_FORMATS[time_format])
    table = recalculate(path, tmp_path)

    expected = distribute(
        [Result(name=name, seconds=seconds) for name, seconds in PROTOCOL],
        fund_from_entries(ENTRIES),
    )

    rows = table[HEADER_ROW : HEADER_ROW + len(PROTOCOL)]
    assert [money(row[4]) for row in rows] == [int(seconds) for _, seconds in PROTOCOL]
    assert [money(row[10]) for row in rows] == [payout.amount for payout in expected.payouts]
    assert [money(row[9]) for row in rows] == [int(payout.raw) for payout in expected.payouts]
    assert [row[8] == "1" for row in rows[:-1]] == [step.funded for step in expected.steps]
    assert money(cell(table, PAID_CELL)) == expected.total_paid
    assert money(cell(table, REMAINDER_CELL)) == expected.remainder
    assert money(cell(table, FUND_CELL)) == expected.fund
    assert cell(table, WARNING_CELL) == ""


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice не установлен")
def test_protocol_in_the_wrong_order_is_flagged(workbook_path: Path, tmp_path: Path) -> None:
    path = tmp_path / "shuffled.xlsx"
    shutil.copy(workbook_path, path)
    book = load_workbook(path)
    sheet = book["Мужчины"]
    for index, (name, seconds) in enumerate(sorted(PROTOCOL, key=lambda row: -row[1])):
        row = FIRST_DATA_ROW + index
        sheet[f"C{row}"] = name
        sheet[f"D{row}"] = seconds
    sheet[ENTRIES_CELL] = ENTRIES
    book.save(path)

    table = recalculate(path, tmp_path)
    assert cell(table, WARNING_CELL) == OUT_OF_ORDER_WARNING

    payouts = [money(row[10]) for row in table[HEADER_ROW : HEADER_ROW + len(PROTOCOL)]]
    assert all(amount >= 0 for amount in payouts), payouts


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice не установлен")
def test_rider_without_a_result_pays_but_does_not_race(workbook_path: Path, tmp_path: Path) -> None:
    path = tmp_path / "dnf.xlsx"
    shutil.copy(workbook_path, path)
    book = load_workbook(path)
    sheet = book["Мужчины"]
    for index, (name, seconds) in enumerate(PROTOCOL):
        row = FIRST_DATA_ROW + index
        sheet[f"C{row}"] = name
        sheet[f"D{row}"] = seconds
    dnf_row = FIRST_DATA_ROW + len(PROTOCOL)
    sheet[f"C{dnf_row}"] = "Сошедший"
    book.save(path)

    table = recalculate(path, tmp_path)
    assert money(cell(table, ENTRIES_CELL)) == len(PROTOCOL) + 1, "сошедший тоже платил взнос"
    dnf = table[dnf_row - 1]
    assert dnf[0] == "", "сошедшему место не полагается"
    assert dnf[4] == "", "и результата у него нет"


FRACTIONAL_PROTOCOL: tuple[tuple[str, str, float], ...] = (
    ("Kazantsev Ilya", "00:29:26.1", 1766.1),
    ("Троицкий Сергей", "00:29:46.9", 1786.9),
    ("Бижан Ғибадат", "00:29:53.4", 1793.4),
    ("Chernykh Denis", "00:30:21.9", 1821.9),
)


@pytest.mark.skipif(shutil.which("soffice") is None, reason="LibreOffice не установлен")
def test_tenths_of_a_second_survive_the_protocol(workbook_path: Path, tmp_path: Path) -> None:
    # Протокол гонки печатает время с десятыми: «00:29:26.1». Раньше на них
    # разбор спотыкался и вся таблица оставалась пустой.
    path = tmp_path / "tenths.xlsx"
    shutil.copy(workbook_path, path)
    book = load_workbook(path)
    sheet = book["Мужчины"]
    for index, (name, shown, _) in enumerate(FRACTIONAL_PROTOCOL):
        row = FIRST_DATA_ROW + index
        sheet[f"C{row}"] = name
        sheet[f"D{row}"] = shown
    book.save(path)

    table = recalculate(path, tmp_path)
    expected = distribute(
        [Result(name=name, seconds=seconds) for name, _, seconds in FRACTIONAL_PROTOCOL],
        fund_from_entries(len(FRACTIONAL_PROTOCOL)),
    )

    def number(text: str) -> float:
        return float(text.replace("\xa0", "").replace(" ", "").replace(",", ".") or 0)

    rows = table[HEADER_ROW : HEADER_ROW + len(FRACTIONAL_PROTOCOL)]
    assert [number(row[4]) for row in rows] == [seconds for _, _, seconds in FRACTIONAL_PROTOCOL]
    assert [number(row[10]) for row in rows] == [payout.amount for payout in expected.payouts]
    assert number(cell(table, PAID_CELL)) == expected.total_paid
